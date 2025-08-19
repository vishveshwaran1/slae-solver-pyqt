from PyQt6.QtWidgets import QListWidgetItem
import numpy as np
import openpyxl
import ast


def subscript(number):
    subscript_map = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
    return str(number).translate(subscript_map)

def int_to_roman(num):
    roman_dict = {
    50: "L",
    49: "XLIX",
    48: "XLVIII",
    47: "XLVII",
    46: "XLVI",
    45: "XLV",
    44: "XLIV",
    43: "XLIII",
    42: "XLII",
    41: "XLI",
    40: "XL",
    39: "XXXIX",
    38: "XXXVIII",
    37: "XXXVII",
    36: "XXXVI",
    35: "XXXV",
    34: "XXXIV",
    33: "XXXIII",
    32: "XXXII",
    31: "XXXI",
    30: "XXX",
    29: "XXIX",
    28: "XXVIII",
    27: "XXVII",
    26: "XXVI",
    25: "XXV",
    24: "XXIV",
    23: "XXIII",
    22: "XXII",
    21: "XXI",
    20: "XX",
    19: "XIX",
    18: "XVIII",
    17: "XVII",
    16: "XVI",
    15: "XV",
    14: "XIV",
    13: "XIII",
    12: "XII",
    11: "XI",
    10: "X",
    9: "IX",
    8: "VIII",
    7: "VII",
    6: "VI",
    5: "V",
    4: "IV",
    3: "III",
    2: "II",
    1: "I"
}

    return roman_dict[num]




class GaussJordanElimination:
    def __init__(self, matrix, B, output_widget):
        self.matrix = np.array(matrix, dtype=float)
        self.B = np.array(B, dtype=float)
        self.Nomer_deistia = 0
        self.output_widget = output_widget
        self.excel_data = []

        
        self.append_text(f"Решение СЛАУ методом Гаусса-Жордана. Матрица: \n{self.matrix}. \nСвободные члены: \n{self.B}")
        
        augmented_matrix = np.hstack((self.matrix, self.B.reshape(-1, 1)))
        self.append_text_for_excel(f"Решение СЛАУ методом Гаусса-Жордана. Матрица:", f"{augmented_matrix.tolist()}")


        
        try:
            self.solve()
        except Exception as e:
            self.append_text(f"Произошла ошибка: {e}")

    def solve(self):
        for i in range(self.matrix.shape[1]): 
            if np.all(self.matrix[:, i] == 0): 
                TextAppend = f"Ответ: Все коэффициенты столбца {int_to_roman(i+1)} (переменная x{subscript(i+1)}) одновременно равны нулю. \nЭто недопустимое условие."
                self.append_text(TextAppend) 
                self.append_text_for_excel(TextAppend, " ")
                return
            
        augmented_matrix = np.hstack((self.matrix, self.B.reshape(-1, 1)))
      
        col_per = augmented_matrix.shape[1] - 1 # количесвто переменных
        n = augmented_matrix.shape[0] # количесвто строк

        for i in range(n):
            if col_per <= i: # Не равное кол строк и переменных
                break
            else:
                TextAppend = f"{self.Vivod_Nomer_deistia_i()} Нормализация строки {int_to_roman(i+1)}: делим строку {int_to_roman(i+1)} на {augmented_matrix[i, i]}"
                self.append_text(TextAppend) 
                self.append_text_for_excel(TextAppend, " ")
                
                augmented_matrix[i] = augmented_matrix[i] / augmented_matrix[i, i]
                
                TextAppend = f"Матрица после нормализации строки {int_to_roman(i+1)}:"
                self.append_text(TextAppend + f"\n{augmented_matrix}") 
                self.append_text_for_excel(TextAppend, f"{augmented_matrix.tolist()}")


                #####
                for j in range(n):
                    if i != j:
                        TextAppend = f"{self.Vivod_Nomer_deistia_i()} Обнуление элемента ({augmented_matrix[j,i]}) в строке {int_to_roman(j+1)}: (вычитаем) cтрока {int_to_roman(j+1)} - ({augmented_matrix[j, i]}) × строку {int_to_roman(i+1)}"
                        self.append_text(TextAppend)
                        self.append_text_for_excel(TextAppend, " ")
                        augmented_matrix[j] = augmented_matrix[j] - augmented_matrix[j, i] * augmented_matrix[i]
                        self.append_text(f"Матрица после обнуления элемента:\n{augmented_matrix}")
                        self.append_text_for_excel(f"Матрица после обнуления элемента:", f"{augmented_matrix.tolist()}")
            
            
                
                
        # if n != augmented_matrix.shape[0]:
            
        augmented_matrix = np.round(augmented_matrix, 10)
        self.Otvet = (augmented_matrix[:, -1])[:col_per]
        
        if col_per < n: # Если строк больше чем переменных
            Nomer_nesoshedsheisa_row = col_per+1
            for i in range(col_per, n): # Проверяем оставшиеся строки на соответствие ответу
                if augmented_matrix[i, -1] != 0:
                    self.append_text(f"Ответ: Равенство {Nomer_nesoshedsheisa_row} не является верным. Система не имеет решения.")
                    self.append_text_for_excel(f"Ответ: Равенство {Nomer_nesoshedsheisa_row} не является верным. Система не имеет решения.", " ")
                    return
                Nomer_nesoshedsheisa_row = Nomer_nesoshedsheisa_row + 1
        if col_per > n: # Если переменных больше чем строк
            final_otv = ""
            for i in range(n): 
                final_otv = final_otv + f"x{subscript(i+1)} = "
                nomer_x = n + 1
                otvet = f"{self.Otvet[i]}"
                for j in range(n, col_per): # По всем кроме свободным (последнего)
                    otvet = otvet + f"{-1 * augmented_matrix[i, j]}x{subscript(nomer_x)} "
                    nomer_x = nomer_x + 1
                otvet = otvet + "\n"
                final_otv = final_otv + otvet 
            self.Otvet = final_otv.split("\n")
            
            TextAppend = f"Ответ: {', '.join(map(str, self.Otvet))}"
            self.append_text(TextAppend)
            self.append_text_for_excel(TextAppend, " ")
            return
                
        TextAppend = f"Ответ: {', '.join(map(str, self.Otvet))}"
        self.append_text(TextAppend) # Вывод Кортежа как str
        self.append_text_for_excel(TextAppend, " ")

    def Vivod_Nomer_deistia_i(self):
        self.Nomer_deistia += 1
        return f"{self.Nomer_deistia}."

    def append_text(self, text):
        print(text)
        self.output_widget.addItem(QListWidgetItem(text))
        
    def append_text_for_excel(self, text, matrix):
        self.excel_data.append(text)
        self.excel_data.append(matrix)
    
    
    
    def save_to_excel(self, excel_file):


        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Результаты"

        row_index = 1
        
       

        for text in self.excel_data:
            # Попытка преобразовать строку в массив NumPy
            otstup_sleva = len(self.matrix[0])+1 

            try:
                if isinstance(text, str):
                    parsed = ast.literal_eval(text)  # Преобразуем строку в объект Python
                    if isinstance(parsed, list) and all(isinstance(row, list) for row in parsed):
                        
                        text = np.array(parsed)  # Преобразуем в NumPy массив

                if isinstance(text, np.ndarray):  # Проверяем, является ли это матрицей

 
                     for row in text:  # Проходим по строкам матрицы
                        for col_index, value in enumerate(row, start=1):
                            sheet.cell(row=row_index, column=col_index, value=value)
                        row_index += 1
                     #
                     
                     ot_row = row_index - len(self.B)
                     for _ in range(len(self.B)): 
                         A = sheet.cell(row=ot_row, column=otstup_sleva)
                         yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                         A.fill = yellow_fill  
                         ot_row = ot_row + 1
                     #   
                else:  # Если это не матрица, записываем как обычный текст
                    sheet.cell(row=row_index, column=1, value=text)
                    row_index += 1

            except (ValueError, SyntaxError):
                # Если преобразование не удалось, записываем строку как есть
                sheet.cell(row=row_index, column=1, value=text)
                row_index += 1

        workbook.save(excel_file)
        print(f"Данные сохранены в файл {excel_file}")




