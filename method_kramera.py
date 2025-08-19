from PyQt6.QtWidgets import QListWidgetItem
import numpy as np
import openpyxl
import ast  

def int_to_roman(num):
    roman_dict = {
    50: "L",
    49: "XLIX",
    48: "XLVIII",
    47: "XLVII",
    46: "XLVI",
    45: "XLV",
    44: "XLIV",
    43: "XLIII",
    42: "XLII",
    41: "XLI",
    40: "XL",
    39: "XXXIX",
    38: "XXXVIII",
    37: "XXXVII",
    36: "XXXVI",
    35: "XXXV",
    34: "XXXIV",
    33: "XXXIII",
    32: "XXXII",
    31: "XXXI",
    30: "XXX",
    29: "XXIX",
    28: "XXVIII",
    27: "XXVII",
    26: "XXVI",
    25: "XXV",
    24: "XXIV",
    23: "XXIII",
    22: "XXII",
    21: "XXI",
    20: "XX",
    19: "XIX",
    18: "XVIII",
    17: "XVII",
    16: "XVI",
    15: "XV",
    14: "XIV",
    13: "XIII",
    12: "XII",
    11: "XI",
    10: "X",
    9: "IX",
    8: "VIII",
    7: "VII",
    6: "VI",
    5: "V",
    4: "IV",
    3: "III",
    2: "II",
    1: "I"
}

    return roman_dict[num]

class MethodKramera:
    def __init__(self, matrix, B, output_widget):
        self.matrix = np.array(matrix)
        self.B = np.array(B)
        self.Nomer_deistia = 0
        self.output_widget = output_widget
        self.excel_data = []

        

        self.append_text(f"Решение СЛАУ методом Крамера.")  
        
        self.append_text(f"Матрица: \n{self.matrix}")
        self.append_text_for_excel("Решение СЛАУ методом Крамера.\nМатрица: \n",f"{self.matrix.tolist()}")
        
        self.append_text(f"Свободные члены: {self.B}")

        try:
            self.opred_matrix = round(float(self.determinant(self.matrix)), 8)
            self.append_text(f"Определитель матрицы: ")
            self.append_text_for_excel("Определитель матрицы:",f"{self.opred_matrix}")
            
            self.append_text(f"{self.opred_matrix}")
            self.matrix_opred_zero = self.uznat_matrix_opred_zero(self.opred_matrix)

            Δ = lambda opred_matrix: "(если Δ=0, то нужно выполнить проверку на наличие решений)" if opred_matrix == 0 else ""
            znack = lambda opred_matrix: "=" if opred_matrix == 0 else "≠"
            
            TextAppend = f"{self.Vivod_Nomer_deistia_i()} Проверка матрицы на нулевой определитель: Δ={self.opred_matrix}{znack(self.opred_matrix)}0 {Δ(self.opred_matrix)}"
            self.append_text(TextAppend)
            

            self.list_opred_so_svobodnimi_chlenami = self.determinant_po_stolb(self.matrix, self.B)
            self.append_text(f"Определители со свободными членами: {self.list_opred_so_svobodnimi_chlenami}")
            
            if self.matrix_opred_zero:
                self.Otvet = self.cheak_na_nalichae_reshenii(self.list_opred_so_svobodnimi_chlenami)
                TextAppend = f"Ответ: {self.Otvet}"
                self.append_text(TextAppend)
                self.append_text_for_excel(TextAppend,f" ")
            else:
                Nomer_deistia = self.Vivod_Nomer_deistia_i()
                TextAppend = f"{Nomer_deistia} Найдем решения (разделим поочередно определители со свободными членами на определитель исходной матрицы):"
                self.append_text(TextAppend)
                self.append_text_for_excel(TextAppend,f" ")
                self.Otvet = tuple(round(float(self.list_opred_so_svobodnimi_chlenami[i]) / self.opred_matrix, 8) for i in range(len(self.list_opred_so_svobodnimi_chlenami)))

                for i in range(len(self.list_opred_so_svobodnimi_chlenami)):
                    TextAppend = f"\t{Nomer_deistia}{i+1} X{i+1}={self.list_opred_so_svobodnimi_chlenami[i]}/{self.opred_matrix}={self.Otvet[i]}"
                    self.append_text(TextAppend)
                    self.append_text_for_excel(TextAppend,f" ")
                TextAppend = f"Ответ: {', '.join(map(str, self.Otvet))}"
                self.append_text(TextAppend) # Вывод Кортежа как str
                self.append_text_for_excel(TextAppend,f" ")                 
        except Exception as e:
            print(f"Произошла ошибка: {e}")
            raise ValueError("Некоректные исходные данные")
    
    def Vivod_Nomer_deistia_i(self):
        self.Nomer_deistia += 1
        return f"{self.Nomer_deistia}."

    def cheak_na_nalichae_reshenii(self, list_opred):
        self.append_text(f"{self.Vivod_Nomer_deistia_i()} Выполнить проверку на наличие решений")
        ot = "Так как Δ=0 и "
        try:
            for value in list_opred:
                if value != 0:
                    ot += f"хотя бы один определитель матрицы со свободными членами не равен 0 (например Δ={value}≠0), то \"Система не имеет решений\""
                    self.append_text(ot)
                    self.append_text_for_excel(ot,f" ")
                    return "Система не имеет решений"
            if max(list_opred) == min(list_opred) == 0:
                ot += "все определители матрицы со свободными членами равны 0, то \"Система имеет бесконечно много решений\""
                self.append_text(ot)
                self.append_text_for_excel(ot,f" ")
                return "Система имеет бесконечно много решений"
        except Exception as e:
            print(f"Произошла ошибка: {e}")

    def determinant(self, matrix):
        TextAppend = f"{self.Vivod_Nomer_deistia_i()} Вычислить определитель для матрицы:"
        self.append_text(TextAppend)
        self.append_text_for_excel(TextAppend,f"{matrix.tolist()}")
        self.append_text(f"{matrix}")
        return np.linalg.det(matrix)
    
    def determinant_po_stolb(self, matrix, list_svobodnix_chlenov):
        TextAppend = f"{self.Vivod_Nomer_deistia_i()} Вычисление определителей матриц со свободными членами."
        self.append_text(TextAppend)
        list_opred_so_svobodnimi_chlenami = []
        for i in range(len(matrix)):
            matrix_so_svobonimi = np.copy(matrix)
            matrix_so_svobonimi[:, i] = list_svobodnix_chlenov
            self.append_text(f"Матрица с подставленными свободными членами в столбец {int_to_roman(i+1)}:")
            self.append_text_for_excel(TextAppend + f"Матрица с подставленными свободными членами в столбец {int_to_roman(i+1)}:",f"{matrix_so_svobonimi.tolist()}")
            
            self.append_text(f"{matrix_so_svobonimi}")
            det = np.linalg.det(matrix_so_svobonimi)
            list_opred_so_svobodnimi_chlenami.append(round(float(det), 8))
            self.append_text(f"Определитель матрицы с подставленными свободными членами в столбец {int_to_roman(i+1)}: {list_opred_so_svobodnimi_chlenami[i]}")
            self.append_text_for_excel(f"Определитель матрицы с подставленными свободными членами в столбец {int_to_roman(i+1)}: {list_opred_so_svobodnimi_chlenami[i]}",f" ")
        return list_opred_so_svobodnimi_chlenami

    def uznat_matrix_opred_zero(self, opred):
        return opred == 0

    def append_text(self, text):
        print(text)
        self.output_widget.addItem(QListWidgetItem(text))
        
    def append_text_for_excel(self, text, matrix):
        self.excel_data.append(text)
        
        self.excel_data.append(matrix)
    

    def save_to_excel(self, excel_file):


        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Результаты"

        row_index = 1
        
        # Свободные члены
        otstup_sleva = len(self.matrix)+1
        for column_da in range(len(self.B)):
            A = sheet.cell(row=column_da+2, column=otstup_sleva, value=self.B[column_da])
            yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            A.fill = yellow_fill

        for text in self.excel_data:
            # Попытка преобразовать строку в массив NumPy


            try:
                if isinstance(text, str):
                    parsed = ast.literal_eval(text)  # Преобразуем строку в объект Python
                    if isinstance(parsed, list) and all(isinstance(row, list) for row in parsed):
                        
                        text = np.array(parsed)  # Преобразуем в NumPy массив

                if isinstance(text, np.ndarray):  # Проверяем, является ли это матрицей
 
                     for row in text:  # Проходим по строкам матрицы
                        for col_index, value in enumerate(row, start=1):
                            sheet.cell(row=row_index, column=col_index, value=value)
                        row_index += 1
                else:  # Если это не матрица, записываем как обычный текст
                    sheet.cell(row=row_index, column=1, value=text)
                    row_index += 1

            except (ValueError, SyntaxError):
                # Если преобразование не удалось, записываем строку как есть
                sheet.cell(row=row_index, column=1, value=text)
                row_index += 1

        workbook.save(excel_file)
        print(f"Данные сохранены в файл {excel_file}")






