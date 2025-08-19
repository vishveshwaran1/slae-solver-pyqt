


def check_sign(value):
    if value >= 0:
        return "+"
    else: return ""

def subscript(number):
    subscript_map = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
    return str(number).translate(subscript_map)

def int_to_roman(num):
    roman_dict = {
    50: "L",
    49: "XLIX",
    48: "XLVIII",
    47: "XLVII",
    46: "XLVI",
    45: "XLV",
    44: "XLIV",
    43: "XLIII",
    42: "XLII",
    41: "XLI",
    40: "XL",
    39: "XXXIX",
    38: "XXXVIII",
    37: "XXXVII",
    36: "XXXVI",
    35: "XXXV",
    34: "XXXIV",
    33: "XXXIII",
    32: "XXXII",
    31: "XXXI",
    30: "XXX",
    29: "XXIX",
    28: "XXVIII",
    27: "XXVII",
    26: "XXVI",
    25: "XXV",
    24: "XXIV",
    23: "XXIII",
    22: "XXII",
    21: "XXI",
    20: "XX",
    19: "XIX",
    18: "XVIII",
    17: "XVII",
    16: "XVI",
    15: "XV",
    14: "XIV",
    13: "XIII",
    12: "XII",
    11: "XI",
    10: "X",
    9: "IX",
    8: "VIII",
    7: "VII",
    6: "VI",
    5: "V",
    4: "IV",
    3: "III",
    2: "II",
    1: "I"
}

    return roman_dict[num]

from re import A
from PyQt6.QtWidgets import QListWidgetItem
import numpy as np
import openpyxl
import ast

class GaussElimination:
    def __init__(self, matrix, B, output_widget):
        self.matrix = np.array(matrix, dtype=float)
        self.B = np.array(B, dtype=float)
        self.Nomer_deistia = 0
        self.output_widget = output_widget
        self.excel_data = []

        self.append_text(f"Решение СЛАУ методом Гаусса. Матрица: \n{self.matrix}. \nСвободные члены: \n{self.B}")
        self.append_text_for_excel(f"Решение СЛАУ методом Гаусса. Матрица:",f"{np.hstack((self.matrix, self.B.reshape(-1, 1))).tolist()}")
        
        try:
            if self.forward_elimination() == "End": return
            
            if self.col_per > self.n: # Если переменных больше чем строк
                    print(self.augmented_matrix)
                
                    final_otv = ""
                    for i in range(self.n): 
                        final_otv = final_otv + f"x{subscript(i+1)} = "
                        nomer_x = self.n
                        otvet = f"{self.augmented_matrix[i,-1]}"
                        for j in range(self.n-1, self.col_per): # По всем кроме свободным (последнего)
                            otvet = otvet + f"{check_sign(-1 * self.augmented_matrix[i, j])}{-1 * self.augmented_matrix[i, j]}x{subscript(nomer_x)} "
                            nomer_x = nomer_x + 1
                        otvet = f"({otvet})" + f"/{self.augmented_matrix[i,i]}" + "\n"
                        final_otv = final_otv + otvet 
                    self.Otvet = final_otv.split("\n")
                    self.append_text(f"Ответ: {', '.join(map(str, self.Otvet))}")
                    self.append_text_for_excel(f"Ответ: {', '.join(map(str, self.Otvet))}", " ")
                    return
            
            self.Otvet = self.backward_substitution()
            
            if self.col_per < self.n: # Если строк больше чем переменных
                    TextAppend = f"{self.Vivod_Nomer_deistia_i()} Проверим оставшиеся строки на соответствие решению"
                    self.append_text(TextAppend) 
                    self.append_text_for_excel(TextAppend, " ")
                    # self.append_text_for_excel(f"Ответ: {', '.join(map(str, self.Otvet))}", " ")
                    Nomer_nesoshedsheisa_row = self.col_per+1
                    for i in range(self.col_per, self.n): # Проверяем оставшиеся строки на соответствие ответу
                        
                        if self.augmented_matrix[i, -1] != self.augmented_matrix[i, -2] * self.Otvet[-1]:
                            TextAppend = f"Строка {Nomer_nesoshedsheisa_row} не является верным, х{subscript(self.col_per)} * {self.augmented_matrix[i, -2]} не равно {self.augmented_matrix[i, -1]} "
                            # self.append_text(f"Строка {Nomer_nesoshedsheisa_row} не является верным, х{subscript(self.col_per)} * {self.augmented_matrix[i, -2]} не равно {self.augmented_matrix[i, -1]} ")
                            self.append_text(TextAppend + "\nОтвет: Система не имеет решения.")
                            self.append_text_for_excel(TextAppend, " ")
                            self.append_text_for_excel("\nОтвет: Система не имеет решения.", " ")
                            return
                        Nomer_nesoshedsheisa_row = Nomer_nesoshedsheisa_row + 1
                        

            
            

            self.append_text(f"Ответ: {', '.join(map(str, self.Otvet))}") # Вывод Кортежа как str
            self.append_text_for_excel(f"Ответ: ",f"{self.Otvet.tolist()}")
        except Exception as e:
            self.append_text(f"Произошла ошибка: {e}")

    def forward_elimination(self):
        augmented_matrix = np.hstack((self.matrix, self.B.reshape(-1, 1)))
        self.n = augmented_matrix.shape[0]
        self.col_per = augmented_matrix.shape[1] - 1
        
        for i in range(self.n):
            if (self.col_per - 1) <= i or i + 1 >= self.n: 
                break
            else:
            
                min_row = i
                for k in range(i + 1, self.n):
                    if abs(augmented_matrix[k, i]) < abs(augmented_matrix[min_row, i]):
                        min_row = k
                if augmented_matrix[min_row, i] == 0:
                
                    TextAppend = f"Ответ: невозможно решить СЛАУ, элемент на главной диагонали равен нулю в столбце {int_to_roman(i+1)}." 
                    self.append_text(TextAppend)
                    self.append_text_for_excel(TextAppend," ")
                    return "End"
                augmented_matrix[[i, min_row]] = augmented_matrix[[min_row, i]]
            
                TextAppend = f"{self.Vivod_Nomer_deistia_i()} Поиск строки с минимальным элементом в столбце {int_to_roman(i+1)}"
                self.append_text(TextAppend)
                self.append_text_for_excel(TextAppend," ")
                if min_row != i:
                    TextAppend = f"{self.Vivod_Nomer_deistia_i()} Обмен строк {int_to_roman(i+1)} и {int_to_roman(min_row+1)}:"
                    self.append_text(TextAppend + f"\n{augmented_matrix}")
                    self.append_text_for_excel(TextAppend, f"{augmented_matrix}")
                else:
                    TextAppend = f"Строка с ближайшим к нулю элементом в текущем столбце это нынешняя строка, обмен строк не нужен."
                    self.append_text(TextAppend)
                    self.append_text_for_excel(TextAppend , " ")
                for k in range(i + 1, self.n):
                    factor = augmented_matrix[k, i] / augmented_matrix[i, i]
                    augmented_matrix[k, i:] -= factor * augmented_matrix[i, i:]
                 
                    TextAppend = f"{self.Vivod_Nomer_deistia_i()} Зануление элемента в строке {int_to_roman(k+1)}, столбец {int_to_roman(i+1)}: (вычитаем) cтрока {int_to_roman(k+1)} - ({factor}) × строку {int_to_roman(i+1)}"
                    self.append_text(TextAppend)
                    self.append_text(f"{augmented_matrix}")
                    self.append_text_for_excel(TextAppend,f"{augmented_matrix.tolist()}")
                

        self.augmented_matrix = augmented_matrix

    def backward_substitution(self):
        x = np.zeros(self.col_per)
        
        col_back_proxodov = self.opred_col_back()
        for i in range(col_back_proxodov - 1, -1, -1):
            if self.augmented_matrix[i, i] == 0:
                TextAppend = "Нулевой элемент на диагонали при обратном ходе"
                self.append_text(TextAppend)
                self.append_text_for_excel(TextAppend, " ")
            
            x[i] = self.augmented_matrix[i, -1] / self.augmented_matrix[i, i]
            
            TextAppend = f"Решаем последнюю строку: x{subscript(i+1)} = {self.augmented_matrix[i, -1]} / {self.augmented_matrix[i,i]} = {x[i]}"
            self.append_text(TextAppend)
            self.append_text_for_excel(TextAppend," ")


            for k in range(i - 1, -1, -1):
                
                
                TextAppend = f"{self.Vivod_Nomer_deistia_i()} Обратный ход для строки {int_to_roman(k+1)}: (вычитаем) {self.augmented_matrix[k,-1]} (из строки {int_to_roman(k+1)}) - ({self.augmented_matrix[k, i]}) из строки ({int_to_roman(k+1)}) × x{subscript(i+1)}"
                self.augmented_matrix[k, -1] -= self.augmented_matrix[k, i] * x[i]
                self.augmented_matrix[k, i] = 0
                self.append_text(TextAppend + "\n" + f"{self.augmented_matrix}")
                self.append_text_for_excel(TextAppend,f"{self.augmented_matrix.tolist()}")
                
        
        x = np.round(x, 10)
        
        


        return x

    def Vivod_Nomer_deistia_i(self):
        self.Nomer_deistia += 1
        return f"{self.Nomer_deistia}."

    def append_text(self, text):
        print(text)
        self.output_widget.addItem(QListWidgetItem(text))
        
    def append_text_for_excel(self, text, matrix):
        self.excel_data.append(text)
        self.excel_data.append(matrix)
    def save_to_excel(self, excel_file):


        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Результаты"

        row_index = 1
        
       

        for text in self.excel_data:
            otstup_sleva = len(self.matrix[0])+1 

            try:
                if isinstance(text, str):
                    parsed = ast.literal_eval(text)  
                    if isinstance(parsed, list) and all(isinstance(row, list) for row in parsed):
                        
                        text = np.array(parsed)  

                if isinstance(text, np.ndarray): 

 
                     for row in text: 
                        for col_index, value in enumerate(row, start=1):
                            sheet.cell(row=row_index, column=col_index, value=value)
                        row_index += 1
                     
                     ot_row = row_index - len(self.B)
                     for _ in range(len(self.B)): 
                         A = sheet.cell(row=ot_row, column=otstup_sleva)
                         yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                         A.fill = yellow_fill  
                         ot_row = ot_row + 1  
                else:  
                    sheet.cell(row=row_index, column=1, value=text)
                    row_index += 1

            except (ValueError, SyntaxError):
                sheet.cell(row=row_index, column=1, value=text)
                row_index += 1

        workbook.save(excel_file)
        print(f"Данные сохранены в файл {excel_file}")

    def opred_col_back(self):
        print("opred_col_back")
        return min(self.n, self.col_per)